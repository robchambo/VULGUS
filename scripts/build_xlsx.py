#!/usr/bin/env python3
"""Build VULGUS_Dictionary_v0.2.xlsx from the canonical JSON lexicon.

Source of truth: dictionary/vulgus_lexicon.json.
Output: dictionary/VULGUS_Dictionary_v0.2.xlsx.
The xlsx is a generated review artefact; do not hand-edit.
"""
from __future__ import annotations
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
JSON_IN = os.path.join(REPO, 'dictionary', 'vulgus_lexicon.json')
XLSX_OUT = os.path.join(REPO, 'dictionary', 'VULGUS_Dictionary_v0.2.xlsx')

DICTIONARY_HEADERS = [
    'ID', 'Word', 'Variants', 'Length', 'Is Single Word', 'Part of Speech',
    'Definition', 'Etymology Note', 'First Attested', 'Severity (1-5)', 'Register',
    'Region', 'Era', 'Semantic Type', 'Tags', 'Family', 'Wordle Eligible',
    'License', 'Puzzle Difficulty (Y/B/R/K)', 'Ship (YES/NO/MAYBE)',
    'Reviewer 1', 'Reviewer 2', 'Date Approved', 'Source', 'Notes',
]

HEADER_FILL = PatternFill('solid', fgColor='1b1b1b')
HEADER_FONT = Font(bold=True, color='f9f9f9')

def _write_header(ws, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 24

def _readme(ws):
    lines = [
        ('VULGUS Lexicon v0.2', True),
        ('Generated from dictionary/vulgus_lexicon.json — do not hand-edit.', False),
        ('Source of truth is the JSON file; re-run scripts/build_xlsx.py after changes.', False),
        ('', False),
        ('Sheets:', True),
        ('  Dictionary  — one row per word with all metadata', False),
        ('  Themes      — puzzle-friendly thematic recipes', False),
        ('  Groups      — approved 4-word groups (puzzle usage record)', False),
        ('  Sources     — citation sources for etymologies', False),
        ('  Categories  — legacy category labels (for humans)', False),
        ('  Severity    — severity scale 1-5', False),
        ('  Regions     — region codes', False),
        ('  SemanticTypes — semantic type codes', False),
        ('  Stats       — dictionary statistics', False),
    ]
    for row, (text, bold) in enumerate(lines, 1):
        c = ws.cell(row=row, column=1, value=text)
        if bold:
            c.font = Font(bold=True)
    ws.column_dimensions['A'].width = 80

def _dictionary(ws, entries):
    _write_header(ws, DICTIONARY_HEADERS)
    for i, e in enumerate(entries, 2):
        ws.cell(row=i, column=1, value=e['id'])
        ws.cell(row=i, column=2, value=e['word'])
        ws.cell(row=i, column=3, value='|'.join(e['variants']) if e['variants'] else '')
        ws.cell(row=i, column=4, value=e['length'])
        ws.cell(row=i, column=5, value='TRUE' if e['is_single_word'] else 'FALSE')
        ws.cell(row=i, column=6, value=e['part_of_speech'])
        ws.cell(row=i, column=7, value=e['definition'])
        ws.cell(row=i, column=8, value=e['etymology'])
        ws.cell(row=i, column=9, value=e['first_attested'] or '')
        ws.cell(row=i, column=10, value=e['severity'])
        ws.cell(row=i, column=11, value=e['register'])
        ws.cell(row=i, column=12, value=e['region'])
        ws.cell(row=i, column=13, value=e['era'])
        ws.cell(row=i, column=14, value=e['semantic_type'])
        ws.cell(row=i, column=15, value=', '.join(e['tags']))
        ws.cell(row=i, column=16, value=e['family'] or '')
        ws.cell(row=i, column=17, value='TRUE' if e['wordle_eligible'] else 'FALSE')
        ws.cell(row=i, column=18, value=e['license'])
        ws.cell(row=i, column=19, value=e['puzzle_difficulty'])
        ws.cell(row=i, column=20, value=e['ship'])
        ws.cell(row=i, column=21, value=e['reviewer_1'])
        ws.cell(row=i, column=22, value=e['reviewer_2'])
        ws.cell(row=i, column=23, value=e['date_approved'])
        ws.cell(row=i, column=24, value=e['source'])
        ws.cell(row=i, column=25, value=e['notes'])
    # Column widths (rough)
    widths = [5, 18, 18, 8, 13, 14, 50, 50, 13, 12, 12, 10, 14, 14, 30, 10,
              14, 22, 14, 14, 10, 10, 14, 18, 20]
    for i, w in enumerate(widths, 1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = w

def _themes(ws):
    _write_header(ws, ['Theme ID', 'Display Name', 'Difficulty Hint', 'Region Filter',
                       'Tag Filter', 'Min Pool Size', 'Example Group', 'Notes'])
    for i, w in enumerate([14, 26, 14, 14, 28, 12, 40, 30], 1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = w

def _groups(ws):
    _write_header(ws, ['Group ID', 'Theme ID', 'Tile 1', 'Tile 2', 'Tile 3', 'Tile 4',
                       'First Used In Puzzle', 'Used In Puzzles', 'Reviewer', 'Notes'])
    for i, w in enumerate([10, 14, 14, 14, 14, 14, 14, 20, 10, 30], 1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = w

def _sources(ws):
    _write_header(ws, ['Source ID', 'Name', 'URL', 'Type', 'License', 'Access Date', 'Notes'])
    rows = [
        ('OED', 'Oxford English Dictionary', 'https://www.oed.com', 'dictionary',
         'subscription / fair-use-paraphrase', '2026-04-10',
         'Quote only short excerpts; paraphrase full definitions.'),
        ('Etymonline', 'Online Etymology Dictionary', 'https://www.etymonline.com',
         'etymology', 'fair-use-attribution', '2026-04-10',
         'Attribute in any published material.'),
        ('Greens', "Green's Dictionary of Slang", 'https://greensdictofslang.com',
         'slang-dictionary', 'fair-use-attribution', '2026-04-10',
         'Primary source for slang provenance.'),
        ('Ofcom', 'Ofcom Offensive Language Research',
         'https://www.ofcom.org.uk/research-and-data/tv-radio-and-on-demand/tv-research/offensive-language',
         'severity-rating', 'public', '2026-04-10',
         'Severity bands informed by Ofcom research.'),
        ('Wiktionary', 'Wiktionary', 'https://en.wiktionary.org', 'dictionary',
         'CC BY-SA 3.0', '2026-04-10',
         'Share-alike licence — attribute and keep downstream open.'),
        ('LDNOOBW', 'List of Dirty, Naughty, Obscene and Otherwise Bad Words',
         'https://github.com/LDNOOBW/List-of-Dirty-Naughty-Obscene-and-Otherwise-Bad-Words',
         'wordlist', 'CC0', '2026-04-10',
         'Used as a rejection checklist, not a source of ship words.'),
        ('SurgeAI', 'Surge AI Profanity List', 'https://github.com/surge-ai/profanity',
         'wordlist', 'MIT', '2026-04-10',
         'Used as a rejection checklist, not a source of ship words.'),
    ]
    for i, row in enumerate(rows, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    for i, w in enumerate([12, 36, 42, 18, 30, 14, 50], 1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = w

def _categories(ws):
    _write_header(ws, ['Category Label', 'Canonical Tag(s)', 'Difficulty Hint',
                       'Description', 'Example Words'])
    rows = [
        ('British Swears', 'british', 'Red', 'Everyday UK profanity, mild to medium.',
         'bollocks, arse, bugger, bloody'),
        ('American Swears', 'american', 'Red', 'Everyday US profanity, mild to medium.',
         'damn, heck, jeez, crap'),
        ('Australian Swears', 'australian', 'Blue',
         'AU/NZ vernacular profanity and mild insults.',
         'drongo, galah, bogan, bloody'),
        ('Scottish/Irish', 'celtic', 'Blue', 'Scots and Hiberno-English colour.',
         'bampot, eejit, feckin, numpty'),
        ('Soft Swears (G-rated)', 'soft-swear', 'Yellow',
         'Safe for all audiences.', 'sugar, fudge, shoot, darn'),
        ('Workplace-Safe', 'workplace-safe', 'Yellow',
         'Unlikely to attract an HR email.', 'shucks, phooey, rats, drat'),
        ('Minced Oaths', 'minced-oath', 'Blue',
         'Euphemistic substitutes for religious oaths.',
         'zounds, gadzooks, jeez, gosh'),
        ('Religious Origin', 'religious', 'Blue',
         'Words tracing to religious taboo.',
         'damn, hell, bejesus, crikey'),
        ('Exclamations', 'exclamation', 'Yellow',
         'Interjections used for surprise, frustration, or emphasis.',
         'blimey, crikey, shucks, phooey'),
        ('Intensifiers', 'intensifier', 'Yellow',
         'Words that add emphasis to other words.',
         'bloody, wicked, darn, hellfire'),
        ('Words for Nonsense', 'nonsense', 'Blue',
         'Terms meaning worthless talk or ideas.',
         'poppycock, balderdash, codswallop, hogwash'),
        ('Words for Idiot', 'fool', 'Red',
         'Terms for a foolish or stupid person.',
         'numpty, wally, plonker, muppet'),
        ('Words for Annoying Person', 'annoying-person', 'Red',
         'Insults for someone who irritates.',
         'pogue, wack, trippin, diss'),
        ('Animals That Are Swears', 'animal', 'Blue',
         'Insults drawn from animal names.',
         'cow, pig, jackass, drongo'),
        ('Foods That Are Swears', 'food', 'Yellow',
         'Food-word substitutes for stronger language.',
         'sugar, fudge, crumbs, peanut'),
        ('Anatomy-Based', 'anatomy', 'Red',
         'Words derived from body parts.',
         'bollocks, arse, knackers, bum'),
        ('Archaic', 'archaic', 'Black',
         'Words mostly out of everyday use.',
         'varlet, knave, zounds, gadzooks'),
        ('Shakespearean Insults', 'shakespearean', 'Black',
         "Insults from Shakespeare's time or style.",
         'varlet, knave, rapscallion, mountebank'),
        ('Victorian / Edwardian', 'victorian', 'Black',
         'Words from 19th and early 20th century English.',
         'poppycock, twaddle, humbug, balderdash'),
        ('Eponymous Swears', 'eponymous', 'Red',
         'Words named after a person or place.',
         'bunkum, nimrod, mountebank, crapper'),
        ('Sci-Fi / Fictional Swears', 'fictional-scifi', 'Red',
         'Invented profanity from speculative fiction.',
         'frak, smeg, frell, gorram'),
        ('Rhyming Slang', 'rhyming-slang', 'Red',
         'Cockney-style rhyming substitutions.',
         'berk'),
        ('Scatological', 'scatological', 'Red',
         'Words derived from bodily waste.',
         'crap, crud, crapper'),
    ]
    for i, row in enumerate(rows, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    for i, w in enumerate([28, 20, 14, 46, 40], 1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = w

def _severity(ws):
    _write_header(ws, ['Level', 'Label', 'Ofcom equivalent', 'Ship?', 'Notes'])
    rows = [
        (1, 'Soft / G-rated', 'Not offensive', 'YES',
         'sugar, fudge, fiddlesticks — safe for all audiences'),
        (2, 'Mild', 'Mild', 'YES', 'damn, crap, arse, blimey'),
        (3, 'Medium', 'Medium', 'YES', 'bollocks, bugger, piss, tosser'),
        (4, 'Strong', 'Strong', 'NO (editorial review only)',
         'Not for puzzle tiles; may be referenced in etymology only.'),
        (5, 'Strongest', 'Strongest', 'NO', 'Never shipped.'),
    ]
    for i, row in enumerate(rows, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    for i, w in enumerate([8, 18, 18, 28, 60], 1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = w

def _regions(ws):
    _write_header(ws, ['Code', 'Region', 'Notes'])
    rows = [
        ('UK', 'United Kingdom (general)', 'Default for British English swears'),
        ('SCO', 'Scotland', 'Scots-specific'),
        ('IE', 'Ireland', 'Hiberno-English'),
        ('US', 'United States', ''),
        ('AU', 'Australia', 'Includes NZ when not otherwise tagged'),
        ('ARCHAIC', 'Archaic / historical', 'No single modern region'),
        ('FICTIONAL', 'Fictional / invented', 'From speculative fiction'),
        ('GLOBAL', 'Global English', 'No single region'),
    ]
    for i, row in enumerate(rows, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    for i, w in enumerate([10, 26, 40], 1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = w

def _semantic_types(ws):
    _write_header(ws, ['Type', 'Description'])
    rows = [
        ('anatomy', 'Derived from a body part'),
        ('scatological', 'Derived from bodily waste or functions'),
        ('sexual', 'Sexual reference (non-violent only)'),
        ('religious', 'Blasphemy, oaths, or religious reference'),
        ('insult-mild', 'A mild personal insult'),
        ('exclamation', 'Used as a standalone interjection'),
        ('minced-oath', 'Euphemistic substitute for a stronger oath'),
        ('intensifier', 'Adds emphasis; grammatical modifier'),
        ('nonsense', 'Means nonsense, rubbish, or unreliable talk'),
        ('euphemism', 'Polite substitute for a stronger word'),
        ('fictional', 'Invented profanity from fiction'),
    ]
    for i, row in enumerate(rows, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 60

def _stats(ws, entries):
    _write_header(ws, ['Metric', 'Value'])
    from collections import Counter
    rows = [
        ('Schema version', '0.2'),
        ('Total entries', len(entries)),
        ('Shipped (YES)', sum(1 for e in entries if e['ship'] == 'YES')),
        ('MAYBE', sum(1 for e in entries if e['ship'] == 'MAYBE')),
        ('Single-word entries', sum(1 for e in entries if e['is_single_word'])),
        ('Phrase entries', sum(1 for e in entries if not e['is_single_word'])),
        ('Wordle-eligible', sum(1 for e in entries if e['wordle_eligible'])),
        ('Severity 1', sum(1 for e in entries if e['severity'] == 1)),
        ('Severity 2', sum(1 for e in entries if e['severity'] == 2)),
        ('Severity 3', sum(1 for e in entries if e['severity'] == 3)),
    ]
    for i, row in enumerate(rows, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    region_counter = Counter(e['region'] for e in entries)
    ws.cell(row=len(rows) + 3, column=1, value='By region').font = Font(bold=True)
    for i, (region, count) in enumerate(region_counter.most_common(), len(rows) + 4):
        ws.cell(row=i, column=1, value=region)
        ws.cell(row=i, column=2, value=count)
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 14

def main():
    with open(JSON_IN, encoding='utf-8') as f:
        d = json.load(f)
    entries = d['entries']

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    _readme(wb.create_sheet('README'))
    _dictionary(wb.create_sheet('Dictionary'), entries)
    _themes(wb.create_sheet('Themes'))
    _groups(wb.create_sheet('Groups'))
    _sources(wb.create_sheet('Sources'))
    _categories(wb.create_sheet('Categories'))
    _severity(wb.create_sheet('Severity'))
    _regions(wb.create_sheet('Regions'))
    _semantic_types(wb.create_sheet('SemanticTypes'))
    _stats(wb.create_sheet('Stats'), entries)

    wb.save(XLSX_OUT)
    print(f'wrote {XLSX_OUT}')

if __name__ == '__main__':
    main()
