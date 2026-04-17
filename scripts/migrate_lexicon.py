#!/usr/bin/env python3
"""One-time migration: read v0.1 xlsx + curation CSV, write canonical JSON.

Output: dictionary/vulgus_lexicon.json (schema v0.2).
Curation input: dictionary/curation/definitions.csv (word, definition, part_of_speech, first_attested).
"""
from __future__ import annotations
import csv, json, os, re
from openpyxl import load_workbook

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_IN = os.path.join(REPO, 'dictionary', 'VULGUS_Dictionary_v0.1.xlsx')
CURATION = os.path.join(REPO, 'dictionary', 'curation', 'definitions.csv')
JSON_OUT = os.path.join(REPO, 'dictionary', 'vulgus_lexicon.json')

# Category 1-4 label -> canonical tag(s). Multi-word labels become hyphenated ids.
CATEGORY_TO_TAGS = {
    'British Swears': ['british'],
    'American Swears': ['american'],
    'Australian Swears': ['australian'],
    'Scottish/Irish': ['celtic'],
    'Soft Swears (G-rated)': ['soft-swear'],
    'Workplace-Safe': ['workplace-safe'],
    'Minced Oaths': ['minced-oath'],
    'Religious Origin': ['religious'],
    'Exclamations': ['exclamation'],
    'Intensifiers': ['intensifier'],
    'Words for Nonsense': ['nonsense'],
    'Words for Idiot': ['fool'],
    'Words for Annoying Person': ['annoying-person'],
    'Animals That Are Swears': ['animal'],
    'Foods That Are Swears': ['food'],
    'Anatomy-Based': ['anatomy'],
    'Archaic': ['archaic'],
    'Shakespearean Insults': ['shakespearean', 'archaic'],
    'Victorian / Edwardian': ['victorian'],
    'Eponymous Swears': ['eponymous'],
    'Sci-Fi / Fictional Swears': ['fictional-scifi'],
    'Rhyming Slang': ['rhyming-slang', 'cockney'],
    'Scatological': ['scatological'],
}

REGION_TO_TAG = {
    'UK': ['uk'], 'US': ['us'], 'AU': ['australian'], 'IE': ['ireland'],
    'SCO': ['scotland'], 'ARCHAIC': ['archaic'], 'FICTIONAL': ['fictional-scifi'],
    'GLOBAL': [],
}

# Severity + Semantic Type -> Register
def derive_register(severity: int, semantic: str, era: str) -> str:
    if era == 'archaic' or semantic in ('minced-oath',):
        if semantic == 'minced-oath':
            return 'euphemism'
        return 'archaic'
    if semantic == 'euphemism':
        return 'euphemism'
    if severity >= 3:
        return 'vulgar'
    if severity == 2:
        return 'informal'
    return 'informal'

SEMANTIC_TO_POS_DEFAULT = {
    'exclamation': 'interj', 'minced-oath': 'interj', 'intensifier': 'adv',
    'euphemism': 'interj', 'religious': 'interj',
    'anatomy': 'noun', 'scatological': 'noun', 'nonsense': 'noun',
    'insult-mild': 'noun', 'fictional': 'interj',
}

def derive_tags(row: dict) -> list[str]:
    tags: list[str] = []
    for i in (1, 2, 3, 4):
        label = row.get(f'Category {i}')
        if not label:
            continue
        for t in CATEGORY_TO_TAGS.get(label, []):
            if t not in tags:
                tags.append(t)
    region = row.get('Region') or ''
    for t in REGION_TO_TAG.get(region, []):
        if t not in tags:
            tags.append(t)
    semantic = row.get('Semantic Type') or ''
    # Some semantic types round-trip as tags
    for t in ('anatomy', 'scatological', 'religious'):
        if semantic == t and t not in tags:
            tags.append(t)
    era = row.get('Era') or ''
    if era == 'archaic' and 'archaic' not in tags:
        tags.append('archaic')
    return tags

def year_from_etym(etym: str) -> int | None:
    if not etym:
        return None
    # Prefer explicit century/year mentions
    m = re.search(r'\b(1[0-9]{3}|20[0-2][0-9])\b', etym)
    if m:
        return int(m.group(1))
    m = re.search(r'(\d{1,2})(?:st|nd|rd|th)\s*C\b', etym)
    if m:
        cent = int(m.group(1))
        return (cent - 1) * 100 + 50  # mid-century approximation
    return None

def load_curation() -> dict[str, dict]:
    out = {}
    if not os.path.exists(CURATION):
        return out
    with open(CURATION, encoding='utf-8', newline='') as f:
        for row in csv.DictReader(f):
            out[row['word'].lower()] = row
    return out

def main():
    wb = load_workbook(XLSX_IN, data_only=True)
    ws = wb['Dictionary']
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    data = [dict(zip(hdr, r)) for r in rows[1:] if r[0] is not None]

    curation = load_curation()
    entries = []
    for r in data:
        word = (r['Word'] or '').strip()
        if not word:
            continue
        key = word.lower()
        cur = curation.get(key, {})
        semantic = r.get('Semantic Type') or ''
        severity = r.get('Severity (1-5)') or 0
        era = r.get('Era') or ''
        region = r.get('Region') or ''
        is_single = ' ' not in word and '-' not in word
        length = len(word.replace(' ', '').replace('-', ''))
        pos = (cur.get('part_of_speech') or '').strip() or SEMANTIC_TO_POS_DEFAULT.get(semantic, 'noun')
        if not is_single and not cur.get('part_of_speech'):
            pos = 'phrase'
        tags = derive_tags(r)
        first_attested = cur.get('first_attested') or ''
        if first_attested:
            try:
                first_attested = int(first_attested)
            except ValueError:
                first_attested = None
        else:
            first_attested = year_from_etym(r.get('Etymology Note') or '')
        register = derive_register(severity, semantic, era)
        ship = (r.get('Ship (YES/NO/MAYBE)') or '').upper()
        wordle_eligible = (
            is_single and 4 <= length <= 7 and severity <= 3 and ship == 'YES'
            and pos != 'phrase'
        )
        entries.append({
            'id': int(r['ID']),
            'word': word,
            'variants': (r.get('Variants') or '').split('|') if r.get('Variants') else [],
            'length': length,
            'is_single_word': is_single,
            'part_of_speech': pos,
            'definition': (cur.get('definition') or '').strip(),
            'etymology': (r.get('Etymology Note') or '').strip(),
            'first_attested': first_attested,
            'severity': severity,
            'region': region,
            'era': era,
            'semantic_type': semantic,
            'tags': tags,
            'register': register,
            'family': (cur.get('family') or '').strip() or None,
            'wordle_eligible': wordle_eligible,
            'license': 'fair-use-attribution',
            'puzzle_difficulty': r.get('Puzzle Difficulty (Y/B/R/K)') or '',
            'ship': ship,
            'reviewer_1': r.get('Reviewer 1') or '',
            'reviewer_2': r.get('Reviewer 2') or '',
            'date_approved': str(r.get('Date Approved') or ''),
            'source': r.get('Source') or '',
            'notes': r.get('Notes') or '',
        })

    out = {
        'schema_version': '0.2',
        'generated_from': 'VULGUS_Dictionary_v0.1.xlsx + curation/definitions.csv',
        'entry_count': len(entries),
        'entries': entries,
    }
    os.makedirs(os.path.dirname(JSON_OUT), exist_ok=True)
    with open(JSON_OUT, 'w', encoding='utf-8') as f:
        json.dump(out, f, indent=2, ensure_ascii=False)
    print(f'wrote {JSON_OUT} with {len(entries)} entries')
    missing_def = [e['word'] for e in entries if not e['definition']]
    if missing_def:
        print(f'  {len(missing_def)} entries missing definition (fill in curation/definitions.csv)')

if __name__ == '__main__':
    main()
